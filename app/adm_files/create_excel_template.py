""" 
Código para criar um template de Excel utilizando a biblioteca openpyxl.
Code by: Marco Antônio Samuelsson
Qualquer alteração ou cópia deste código deve ser aprovada pela criador!
"""
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#   Importação de bibliotecas para o funcionamento do código
#       openpyxl: para a criação e manipulação dos dados das planilhas
#       openpyxl.styles: fornece classes para estilizar células do Excel, como Font (fonte), PatternFill (preenchimento) e Alignment (alinhamento)
#           Font: permite definir o estilo da fonte (negrito, cor, tamanho, etc.) das células.
#           PatternFill: permite definir o preenchimento de fundo das células (cor, padrão, etc.).
#           Alignment: permite definir o alinhamento do texto dentro das células (centralizado, à esquerda, quebra de linha, etc.).
#       openpyxl.utils: fornece utilitários para manipulação de planilhas, como conversão de índices de coluna para letras (get_column_letter), validação de referências de célula, etc.
#       platform: fornece informações sobre o sistema operacional em execução, útil para abrir o arquivo Excel de forma compatível com Windows, macOS ou Linux.
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import openpyxl.utils
import subprocess
import platform

#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#   Classe para criação de templates Excel
#   Esta classe permite criar um arquivo Excel com múltiplas planilhas, cabeçalhos personalizados e linhas de dados.
#   Principais Métodos:
#       - add_sheet: Adiciona uma nova planilha ao arquivo Excel.
#       - add_head: Adiciona um cabeçalho à planilha especificada.
#       - add_line: Adiciona uma linha de dados à planilha especificada.
#       - save: Salva o arquivo Excel.
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
class Create_Excel():
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#   Inicializa a classe e cria um novo arquivo Excel
#   Parâmetros:
#       - file_name: Nome do arquivo Excel a ser criado.
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
    def __init__(self, file_name: str):
        self.file_name = file_name
        # Cria um novo arquivo Excel
        self.workbook = openpyxl.Workbook()
        # Dicionários para armazenar as planilhas e as larguras das colunas
        self.sheets = {}
        self.max_col_widths = {}  # Armazena a largura máxima por coluna
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#   Método 'add_sheet': Adiciona uma nova planilha ao arquivo Excel
#   Parâmetros:
#       - sheet_name: Nome da planilha a ser adicionada.
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
    def add_sheet(self, sheet_name: str):
        # Verifica se a planilha já existe
        if sheet_name in self.sheets:
            raise ValueError(f"The sheet '{sheet_name}' already exists.")
        # Verifica se é a primeira planilha
        # Se for, renomeia a planilha padrão
        if len(self.workbook.sheetnames) == 1 and self.workbook.active.title == "Sheet":
            sheet = self.workbook.active
            sheet.title = sheet_name
        else:
            sheet = self.workbook.create_sheet(title=sheet_name)
        # Armazena a nova planilha
        self.sheets[sheet_name] = sheet
        # Armazena a largura máxima das colunas
        self.max_col_widths[sheet_name] = {}
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#   Método 'add_head': Adiciona um cabeçalho à planilha especificada.
#   Parâmetros:
#       - sheet_name: Nome da planilha onde o cabeçalho será adicionado.
#       - headers: Lista de strings representando os cabeçalhos das colunas.
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
    def add_head(self, sheet_name: str, headers: list):
        sheet = self.sheets[sheet_name]
        # Muda a cor de fundo e a fonte dos cabeçalhos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        # Para cada cabeçalho
        for col, cabecalho in enumerate(headers, start=1):
            # Cria a célula do cabeçalho
            cell = sheet.cell(row=1, column=col, value=cabecalho)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True)
            self.max_col_widths[sheet_name][col] = len(str(cabecalho))
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#   Método 'add_line': Adiciona uma linha de dados à planilha especificada.
#   Parâmetros:
#       - sheet_name: Nome da planilha onde a linha será adicionada.
#       - datas: Lista de valores a serem adicionados na linha.
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
    def add_line(self, sheet_name: str, datas: list):
        sheet = self.sheets[sheet_name]
        linha_atual = sheet.max_row + 1
        for col, valor in enumerate(datas, start=1):
            cell = sheet.cell(row=linha_atual, column=col, value=valor)
            cell.alignment = Alignment(wrap_text=True)
            atual = self.max_col_widths[sheet_name].get(col, 0)
            novo = max(atual, len(str(valor)))
            self.max_col_widths[sheet_name][col] = novo
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#   Método 'adjust_column_width': Ajusta a largura das colunas com base no conteúdo.
#   Parâmetros:
#       Nenhum
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
    def adjust_column_width(self):
        for sheet_name, col_widths in self.max_col_widths.items():
            sheet = self.sheets[sheet_name]
            for col, width in col_widths.items():
                letra_coluna = openpyxl.utils.get_column_letter(col)
                sheet.column_dimensions[letra_coluna].width = min(50, width + 5)
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#   Método 'save': Salva o arquivo Excel.
#   Parâmetros:
#       Nenhum
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
    def save(self):
        self.adjust_column_width()
        self.workbook.save(self.file_name)
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#   Método 'open_excel_file': Abre o arquivo Excel depois de criado.
#   Parâmetros:
#       Nenhum
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
    def open_excel_file(self):
        system = platform.system()
        try:
            if system == "Windows":
                subprocess.run(["start", "", self.file_name], shell=True)
            elif system == "Darwin":  # macOS
                subprocess.run(["open", self.file_name])
            else:  # Linux
                subprocess.run(["xdg-open", self.file_name])
        except Exception as e:
            return e